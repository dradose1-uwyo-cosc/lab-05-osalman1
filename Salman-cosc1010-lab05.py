import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.Workbook()
sheet = wb.active

for col in range(1, 21):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 3
for row in range(1, 21):
    sheet.row_dimensions[row].height = 18

pixels = {
    (1, 1): "FF0000",
    (1, 2): "FF0000",
    (2, 1): "00FF00",
    (2, 2): "00FF00",
    (3, 1): "0000FF",
    (3, 2): "0000FF",
    (4, 1): "FFFF00",
    (4, 2): "FFFF00",
    (5, 1): "FFA500",
    (5, 2): "FFA500",
}

for (row, col), color in pixels.items():
    cell = sheet.cell(row=row, column=col)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

wb.save("YourLastName_HW5.xlsx")